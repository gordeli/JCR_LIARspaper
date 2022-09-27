# This code loads the LIWC results for the reviews

# import timing
import openpyxl
import nltk
from nltk.corpus import wordnet as wn
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.tag.mapping import tagset_mapping, map_tag
from nltk.stem import WordNetLemmatizer

import sqlite3
import shutil # we use this library to create a copy of a file (in this case to duplicate the database
# so that we can loop over one instance while editing the other)
# Establish a SQLite connection to a database named 'Liars4.sqlite':
conn = sqlite3.connect('Liars7_www_Rel_sent.sqlite')
# Get the cursor, which is used to traverse the database, line by line
cur = conn.cursor()
# Then we duplicate thedatabase, so that one can loop and edit it at the same time
# and 'open' the other 'instance' of the same database
shutil.copyfile('Liars7_www_Rel_sent.sqlite', 'Liars7_w.sqlite')
conn_w = sqlite3.connect('Liars7_w.sqlite')
cur_w = conn_w.cursor()

# Create columns for LIWC data:
try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD WC_LIWC INTEGER DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'WC_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Analytic_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Analytic_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Clout_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Clout_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Authentic_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Authentic_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Tone_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Tone_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD WPS_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'WPS_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Sixltr_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Sixltr_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Dic_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Dic_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD function_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'function_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD pronoun_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'pronoun_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD ppron_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'ppron_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD i_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'i_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD we_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'we_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD you_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'you_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD shehe_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'shehe_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD they_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'they_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD ipron_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'ipron_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD article_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'article_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD prep_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'prep_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD auxverb_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'auxverb_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD adverb_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'adverb_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD conj_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'conj_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD negate_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'negate_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD verb_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'verb_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD adj_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'adj_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD compare_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'compare_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD interrog_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'interrog_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD number_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'number_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD quant_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'quant_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD affect_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'affect_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD posemo_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'posemo_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD negemo_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'negemo_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD anx_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'anx_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD anger_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'anger_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD sad_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'sad_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD social_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'social_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD family_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'family_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD friend_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'friend_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD female_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'female_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD male_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'male_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD cogproc_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'cogproc_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD insight_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'insight_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD cause_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'cause_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD discrep_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'discrep_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD tentat_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'tentat_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD certain_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'certain_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD differ_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'differ_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD percept_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'percept_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD see_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'see_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD hear_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'hear_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD feel_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'feel_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD bio_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'bio_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD body_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'body_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD health_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'health_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD sexual_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'sexual_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD ingest_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'ingest_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD drives_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'drives_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD affiliation_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'affiliation_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD achieve_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'achieve_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD power_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'power_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD reward_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'reward_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD risk_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'risk_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD focuspast_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'focuspast_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD focuspresent_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'focuspresent_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD focusfuture_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'focusfuture_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD relativ_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'relativ_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD motion_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'motion_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD space_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'space_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD time_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'time_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD work_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'work_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD leisure_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'leisure_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD home_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'home_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD money_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'money_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD relig_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'relig_LIWC' exists already''')
    pass # handle the error

try: # cleaned
    cur_w.execute('''ALTER TABLE Sentences_raw ADD death_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'death_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD informal_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'informal_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD swear_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'swear_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD netspeak_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'netspeak_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD assent_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'assent_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD nonflu_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'nonflu_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD filler_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'filler_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD AllPunc_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'AllPunc_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Period_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Period_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Comma_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Comma_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Colon_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Colon_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD SemiC_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'SemiC_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD QMark_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'QMark_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Exclam_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Exclam_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Dash_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Dash_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Quote_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Quote_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Apostro_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Apostro_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD Parenth_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'Parenth_LIWC' exists already''')
    pass # handle the error

try: # raw
    cur_w.execute('''ALTER TABLE Sentences_raw ADD OtherP_LIWC REAL DEFAULT 0''') # DEFAULT 0 was removed from the sql string
except:
    print('''The column 'OtherP_LIWC' exists already''')
    pass # handle the error

# First, let's move brysbaert into SQL:
#Input the name of the excel file to be converted into a SQL database
# name = input("Enter excel file name:")
# if len(name) < 1 : name = "Concreteness_ratings_Brysbaert_et_al_BRM.xlsx"

name = "sentences_raw_LIWC2015_results.xlsx"
name_raw = "sentences_raw_LIWC2015_results.xlsx"

# Open the workbook and define the worksheet:
wb = openpyxl.load_workbook(name, data_only=True)
wb_raw = openpyxl.load_workbook(name_raw)

# We could add input() function to select the sheets we would like
#to convert into the database
sheet = wb.active #selecting the first sheet only
sheet_raw = wb_raw.active #selecting the first sheet only
#sheet = book.sheet_by_name('Organized')
# Create a new table in the database:
# Note: The first line after 'try' statement deletes the table if it already exists.
# This is good at the stage of twicking your code. After the code for importing into
# the database from excel is finalized it is better to remove this: 'DROP TABLE IF EXISTS Reviews;'
# The 'try except else' that follows will generate a message if the table we are
# trying to create already exists.

#Create a For loop to iterate through each row in the XLS file,
#starting at row 3 by default to skip the headers:
for r in range(2, sheet.max_row + 1):
    try: # attempt to extract the content of the relevant cells
        id                  = int(sheet.cell(r,1).value)
        WC                  = int(sheet.cell(r,3).value)
        Analytic            = float(sheet.cell(r,4).value)
        Clout               = float(sheet.cell(r,5).value)
        Authentic           = float(sheet.cell(r,6).value)
        Tone                = float(sheet.cell(r,7).value)
        WPS                 = float(sheet.cell(r,8).value)
        Sixltr              = float(sheet.cell(r,9).value)
        Dic                 = float(sheet.cell(r,10).value)
        function            = float(sheet.cell(r,11).value)
        pronoun             = float(sheet.cell(r,12).value)
        ppron               = float(sheet.cell(r,13).value)
        i                   = float(sheet.cell(r,14).value)
        we                  = float(sheet.cell(r,15).value)
        you                 = float(sheet.cell(r,16).value)
        shehe               = float(sheet.cell(r,17).value)
        they                = float(sheet.cell(r,18).value)
        ipron               = float(sheet.cell(r,19).value)
        article             = float(sheet.cell(r,20).value)
        prep                = float(sheet.cell(r,21).value)
        auxverb             = float(sheet.cell(r,22).value)
        adverb              = float(sheet.cell(r,23).value)
        conj                = float(sheet.cell(r,24).value)
        negate              = float(sheet.cell(r,25).value)
        verb                = float(sheet.cell(r,26).value)
        adj                 = float(sheet.cell(r,27).value)
        compare             = float(sheet.cell(r,28).value)
        interrog            = float(sheet.cell(r,29).value)
        number              = float(sheet.cell(r,30).value)
        quant               = float(sheet.cell(r,31).value)
        affect              = float(sheet.cell(r,32).value)
        posemo              = float(sheet.cell(r,33).value)
        negemo              = float(sheet.cell(r,34).value)
        anx                 = float(sheet.cell(r,35).value)
        anger               = float(sheet.cell(r,36).value)
        sad                 = float(sheet.cell(r,37).value)
        social              = float(sheet.cell(r,38).value)
        family              = float(sheet.cell(r,39).value)
        friend              = float(sheet.cell(r,40).value)
        female              = float(sheet.cell(r,41).value)
        male                = float(sheet.cell(r,42).value)
        cogproc             = float(sheet.cell(r,43).value)
        insight             = float(sheet.cell(r,44).value)
        cause               = float(sheet.cell(r,45).value)
        discrep             = float(sheet.cell(r,46).value)
        tentat              = float(sheet.cell(r,47).value)
        certain             = float(sheet.cell(r,48).value)
        differ              = float(sheet.cell(r,49).value)
        percept             = float(sheet.cell(r,50).value)
        see                 = float(sheet.cell(r,51).value)
        hear                = float(sheet.cell(r,52).value)
        feel                = float(sheet.cell(r,53).value)
        bio                 = float(sheet.cell(r,54).value)
        body                = float(sheet.cell(r,55).value)
        health              = float(sheet.cell(r,56).value)
        sexual              = float(sheet.cell(r,57).value)
        ingest              = float(sheet.cell(r,58).value)
        drives              = float(sheet.cell(r,59).value)
        affiliation         = float(sheet.cell(r,60).value)
        achieve             = float(sheet.cell(r,61).value)
        power               = float(sheet.cell(r,62).value)
        reward              = float(sheet.cell(r,63).value)
        risk                = float(sheet.cell(r,64).value)
        focuspast           = float(sheet.cell(r,65).value)
        focuspresent        = float(sheet.cell(r,66).value)
        focusfuture         = float(sheet.cell(r,67).value)
        relativ             = float(sheet.cell(r,68).value)
        motion              = float(sheet.cell(r,69).value)
        space               = float(sheet.cell(r,70).value)
        time                = float(sheet.cell(r,71).value)
        work                = float(sheet.cell(r,72).value)
        leisure             = float(sheet.cell(r,73).value)
        home                = float(sheet.cell(r,74).value)
        money               = float(sheet.cell(r,75).value)
        relig               = float(sheet.cell(r,76).value)
        death               = float(sheet.cell(r,77).value)
        informal            = float(sheet_raw.cell(r,78).value)
        swear               = float(sheet_raw.cell(r,79).value)
        netspeak            = float(sheet_raw.cell(r,80).value)
        assent              = float(sheet_raw.cell(r,81).value)
        nonflu              = float(sheet_raw.cell(r,82).value)
        filler              = float(sheet_raw.cell(r,83).value)
        AllPunc             = float(sheet_raw.cell(r,84).value)
        Period              = float(sheet_raw.cell(r,85).value)
        Comma               = float(sheet_raw.cell(r,86).value)
        Colon               = float(sheet_raw.cell(r,87).value)
        SemiC               = float(sheet_raw.cell(r,88).value)
        QMark               = float(sheet_raw.cell(r,89).value)
        Exclam              = float(sheet_raw.cell(r,90).value)
        Dash                = float(sheet_raw.cell(r,91).value)
        Quote               = float(sheet_raw.cell(r,92).value)
        Apostro             = float(sheet_raw.cell(r,93).value)
        Parenth             = float(sheet_raw.cell(r,94).value)
        OtherP              = float(sheet_raw.cell(r,95).value)
    except IndexError: #handling the error: print a possible error source
        print('One or several of the cells selected may be out of the table range. Please doublecheck the location of the column from which to parse the data')

    cur_w.execute('''UPDATE Sentences_raw SET (WC_LIWC, Analytic_LIWC, Clout_LIWC, Authentic_LIWC, Tone_LIWC, WPS_LIWC, Sixltr_LIWC, Dic_LIWC, function_LIWC, pronoun_LIWC, ppron_LIWC, i_LIWC, we_LIWC, you_LIWC, shehe_LIWC, they_LIWC, ipron_LIWC, article_LIWC, prep_LIWC, auxverb_LIWC, adverb_LIWC, conj_LIWC, negate_LIWC, verb_LIWC, adj_LIWC, compare_LIWC, interrog_LIWC, number_LIWC, quant_LIWC, affect_LIWC, posemo_LIWC, negemo_LIWC, anx_LIWC, anger_LIWC, sad_LIWC, social_LIWC, family_LIWC, friend_LIWC, female_LIWC, male_LIWC, cogproc_LIWC, insight_LIWC, cause_LIWC, discrep_LIWC, tentat_LIWC, certain_LIWC, differ_LIWC, percept_LIWC, see_LIWC, hear_LIWC, feel_LIWC, bio_LIWC, body_LIWC, health_LIWC, sexual_LIWC, ingest_LIWC, drives_LIWC, affiliation_LIWC, achieve_LIWC, power_LIWC, reward_LIWC, risk_LIWC, focuspast_LIWC, focuspresent_LIWC, focusfuture_LIWC, relativ_LIWC, motion_LIWC, space_LIWC, time_LIWC, work_LIWC, leisure_LIWC, home_LIWC, money_LIWC, relig_LIWC, death_LIWC, informal_LIWC, swear_LIWC, netspeak_LIWC, assent_LIWC, nonflu_LIWC, filler_LIWC, AllPunc_LIWC, Period_LIWC, Comma_LIWC, Colon_LIWC, SemiC_LIWC, QMark_LIWC, Exclam_LIWC, Dash_LIWC, Quote_LIWC, Apostro_LIWC, Parenth_LIWC, OtherP_LIWC) = (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) WHERE id = ?''', (WC, Analytic, Clout, Authentic, Tone, WPS, Sixltr, Dic, function, pronoun, ppron, i, we, you, shehe, they, ipron, article, prep, auxverb, adverb, conj, negate, verb, adj, compare, interrog, number, quant, affect, posemo, negemo, anx, anger, sad, social, family, friend, female, male, cogproc, insight, cause, discrep, tentat, certain, differ, percept, see, hear, feel, bio, body, health, sexual, ingest, drives, affiliation, achieve, power, reward, risk, focuspast, focuspresent, focusfuture, relativ, motion, space, time, work, leisure, home, money, relig, death, informal, swear, netspeak, assent, nonflu, filler, AllPunc, Period, Comma, Colon, SemiC, QMark, Exclam, Dash, Quote, Apostro, Parenth, OtherP, id))

        #cur.execute('SELECT id FROM Product WHERE asin = ? ', (asin, ))
        #Product_id = cur.fetchone()[0]

# Commit the transaction
conn_w.commit()
cur_w.close()
conn_w.close()
cur.close()
conn.close()
shutil.copyfile('Liars7_w.sqlite', 'Liars7_LIWC_sent_fixed.sqlite')
